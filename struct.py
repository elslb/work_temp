# Created by Erich Fernandes, FERNE22, Nov 2022
# Some things in this script can be made into functions for efficiency
# But for easy testing I left them as is
# Because there are only two uses of the same function
# If more sheets are added, create a function around the dataframe and worksheet lines and use that instead
# Would be more efficient use of lines

from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
import calendar
import os
import re
import win32com.client

import time
import glob
import win32com.client as win32
import pandas as pd
import shutil
from sqlalchemy import false
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from xlrd import open_workbook
from xlutils.copy import copy

from pandas.tseries.holiday import USFederalHolidayCalendar

# Removing chained assignment warning
pd.options.mode.chained_assignment = None

uscal = USFederalHolidayCalendar()
user = os.getlogin()

today = datetime.today()
today = datetime(2023, 11, 1)

todaystr = datetime.strftime(today,"%m/%d/%Y %H:%M")
todaystr2 = datetime.strftime(today,"%M%D%Y")
todaystr3 = datetime.strftime(today, "%m/%d")
todaystr4 = datetime.strftime(today,"%Y%m%d")
todaystr5 = datetime.strftime(today, '%Y-%M-%D')
todaystr6 = datetime.strftime(today, '%m/%d/%Y')


todaymonth = datetime.strftime(today, "%m")
# todaymonth = '01'
todaydaymonth = datetime.strftime(today,'%m%d')
todayyear = datetime.strftime(today,'%Y')
# todayyear = '2023'

usholidays = uscal.holidays(start=f'{todayyear}-01-01',end=f'{todayyear}-12-31').to_pydatetime()

currentmonthstart = datetime(year=int(todayyear),month=int(todaymonth),day=1)
nextmonthstart = currentmonthstart + relativedelta(months=+1)

currentmonthlist = pd.date_range(currentmonthstart,nextmonthstart-timedelta(days=1)).to_pydatetime().tolist()

templatefile = rf'\\CIBG-SRV-TOR08\dpss\ged_applications\beta\Structured Notes\Template\Blank calc file.xlsx'
inputdirectory = rf'\\CIBG-SRV-TOR08\dpss\ged_applications\beta\Structured Notes\Input'
outputdirectory = rf'\\CIBG-SRV-TOR08\dpss\ged_applications\beta\Structured Notes\Output'

#inputfile = inputdirectory+rf'\ValuationDateExtract_2022-11-30.xlsx'
inputfile = inputdirectory+rf'\ValuationDateExtract_2023-11-01.xlsx'




print('Obtaining Structures Data')
maindf = pd.read_excel(inputfile, sheet_name = 'Structures')
print('Obtaining Notional Data')
notionaldf = pd.read_excel(inputfile, sheet_name = 'Positional')
print('Obtaining Ticker Data')
tickerdf = pd.read_excel(inputdirectory+rf'\tickers.xlsx')
templatecols = (pd.read_excel(templatefile)).columns

combineddf = maindf.join(notionaldf.set_index('Package Code'),on='Package Code', lsuffix='',rsuffix='_notional')
combineddf.drop_duplicates(subset=['Package Code','Observation Date','Settlement Date','Notional'],keep='first',inplace=True,ignore_index=True)

# combineddf = combineddf.astype({'CUSIP':str})
# combineddf['CUSIP'] = combineddf['CUSIP'].str.strip()

# combineddf = combineddf.astype({'ISM Code':str})
# combineddf['ISM Code'] = combineddf['ISM Code'].str.strip()

# combineddf = combineddf.astype({'Long Name':str})
# combineddf['Long Name'] = combineddf['Long Name'].str.strip()



tickerlist = tickerdf['ticker'].str.rsplit('.',n=1,expand=True)[0]


testcounter = 0

for day in currentmonthlist:
    # print(day)
    if (day not in usholidays) & (day.isoweekday() != 6) & (day.isoweekday() !=7):
        fileloc = outputdirectory+rf'\{datetime.strftime(day,"%b %#d")}.xlsx'
        if testcounter == 40:
            break
        try:
            os.remove( outputdirectory+rf'\{datetime.strftime(day,"%b %#d")}.xlsx')

            os.remove( outputdirectory+rf'\cusip_draft_'+rf'{datetime.strftime(day,"%b %#d")}.csv')
            os.remove( outputdirectory+rf'\isin_draft_'+rf'{datetime.strftime(day,"%b %#d")}.csv')
            os.remove( outputdirectory+rf'\df_'+rf'{datetime.strftime(day,"%b %#d")}.csv')
        except:
            pass
        
        if os.path.exists(fileloc) == False:
            shutil.copy(templatefile,fileloc)
            print(day,'file created.')
        else:
            print(day,'file exists.')   
            
           
        daydf = combineddf.loc[combineddf['Observation Date']==datetime.strftime(day,'%Y-%m-%d')]
        if daydf.empty == False:
            daydf['Long Name'] = daydf['Long Name'].fillna('None')
                    # Adding ISM Code to CUSIP if Structure is Callable
            daydf['CUSIP']=daydf[['CUSIP','ISM Code']].add('\n').sum(axis=1).where(daydf['Long Name'].str.contains('Issuer',case=False),other=daydf['CUSIP']).str.rstrip()

            
            daydf["Autocall Field"] = ""
            # daydf["Paying Interest"] = "=IF(CLOSING VALUE>=BARRIER VALUE,TRUE,FALSE)"
            daydf.loc[daydf['Settlement Date']==daydf['Structure Maturity'],'Autocall Field'] = 'Maturity'
            daydf.loc[daydf['downstrikePayout']==100,'Autocall Field'] = '=IF(CLOSING VALUE >= CALL THRESHOLD VALUE, TRUE FALSE)'
            daydf.loc[((daydf['downstrikePayout']== 0) | (daydf['Long Name'].str.contains('Blackrock', case=False))),'Autocall Field'] = 'N/A'
            daydf.loc[daydf['Long Name'].str.contains('issuer', case=False),'Autocall Field'] = 'Valid IC Date'
            daydf["Interest"] = ""        
        
            cusipdf = daydf[daydf['CUSIP'].str[:2] == '89']
            cusipdf = cusipdf.reset_index(drop=True)
            isindf = daydf[daydf['CUSIP'].str[:2] != '89']
            isindf = isindf.reset_index(drop=True)
            
            
            # Need to add Notional and Principal parsing
            if cusipdf.empty == False:
                # Locating reference assets
                cusiprefassetlist = []
                cusipfinaldaylist = []
                cusipvallist = []
                cusipsettlelist = []
                cusiplist = []
                cusipnotional = []
                cusipmemory = []
                cusipinv = []
                cusipident = []
                cusipauto = []
                cusiptpayingint = []
                cusipinterest = []
                cusiplongname = []


                # Using regex and tickerlist to find tickers in Long Name
                cusiptickers = cusipdf['Long Name'].apply(lambda x: set.intersection(set(re.split('[ /]',x)), set(tickerlist)))
                tickcounter = 0
                    
                for i in range(cusipdf.shape[0]):
                    cusipdf.at[i, 'Paying Interest'] = f"=IF(AA{i+1}>=Y{i+1},TRUE,FALSE)"
                    # cusipdf.at[i, 'Interest'] = = f"=AC{i+1}*D{i+1}/12"


                # print('Printing cusiptickers {}'.format(cusiptickers))
                for tickers in cusiptickers:
                    tickers = list(tickers)
                    # print(tickers)
                    temptickerlist = []
                    if len(tickers) == 0:
                        cusiprefassetlist.append('NoTickerFound')    
                        cusipvallist.append(datetime.strftime(cusipdf['Observation Date'][tickcounter],'%m/%d/%Y'))
                        cusipsettlelist.append(datetime.strftime(cusipdf['Settlement Date'][tickcounter], '%m/%d/%Y'))
                        cusiplist.append(cusipdf['CUSIP'][tickcounter])
                        cusipnotional.append(cusipdf['Notional'][tickcounter])
                        cusipinv.append(cusipdf['inventoryName'][tickcounter])
                        cusipident.append(cusipdf['ident'][tickcounter])
                        cusipfinaldaylist.append([datetime.strftime(cusipdf['Observation Date'][tickcounter], '%m/%d/%Y'),datetime.strftime(cusipdf['Settlement Date'][tickcounter],'%m/%d/%Y')\
                            ,cusipdf['CUSIP'][tickcounter],'NoTickerFound'])
                        cusipauto.append(cusipdf['Autocall Field'][tickcounter])
                        cusiptpayingint.append(cusipdf['Paying Interest'][tickcounter])
                        cusipinterest.append(cusipdf['Interest'][tickcounter])
                        
                        cusiplongname.append(cusipdf['Long Name'][tickcounter])
                        
                        if ('memory' in cusipdf['Long Name'][tickcounter].lower()):
                            cusipmemory.append('Refer to previous month')
                        else:
                            cusipmemory.append('')
                    else:                  
                        for ticker in tickers:     
                            if ticker != 'TD':
                                cusiprefassetlist.append(ticker)
                                cusipvallist.append(datetime.strftime(cusipdf['Observation Date'][tickcounter],'%m/%d/%Y'))
                                cusipsettlelist.append(datetime.strftime(cusipdf['Settlement Date'][tickcounter],'%m/%d/%Y'))
                                cusiplist.append(cusipdf['CUSIP'][tickcounter])
                                cusipnotional.append(cusipdf['Notional'][tickcounter])
                                cusipinv.append(cusipdf['inventoryName'][tickcounter])
                                cusipident.append(cusipdf['ident'][tickcounter])
                                cusipfinaldaylist.append([datetime.strftime(cusipdf['Observation Date'][tickcounter],'%m/%d/%Y'),datetime.strftime(cusipdf['Settlement Date'][tickcounter],'%m/%d/%Y')\
                                    ,cusipdf['CUSIP'][tickcounter],ticker])
                                cusipauto.append(cusipdf['Autocall Field'][tickcounter])
                                cusiptpayingint.append(cusipdf['Paying Interest'][tickcounter])
                                cusipinterest.append(cusipdf['Interest'][tickcounter])
                                cusiplongname.append(cusipdf['Long Name'][tickcounter])
                            
                                
                                if ('memory' in cusipdf['Long Name'][tickcounter].lower()):
                                        cusipmemory.append('Refer to previous month')
                                else:
                                    cusipmemory.append('')

                                
                    tickcounter += 1  
            
            # df = pd.DataFrame(list(zip(cusiprefassetlist,cusipvallist,cusipsettlelist,cusiplist,cusipnotional,cusipinv,cusipident,cusiplongname,cusipauto)),columns =['ref','Observation Date','Settlement Date','Cusip','Notional','inventoryName','ident','Long Name','Autocall'])

            if isindf.empty == False:
            # Locating reference assets
                isinrefassetlist = []
                isinfinaldaylist = []
                isinvallist = []
                isinsettlelist = []
                isinlist = []
                isinnotional = []
                isinmemory = []
                isininv = []
                isinident = []
                isinauto = []
                isintpayingint = []
                isininterest = []
                isinlongname = []
                

                # Using regex and tickerlist to find tickers in Long Name
                isintickers = isindf['Long Name'].apply(lambda x: set.intersection(set(re.split('[ /]',x)), set(tickerlist)))
                tickcounter = 0
                
                for i in range(isindf.shape[0]):
                    isindf.at[i, 'Paying Interest'] = f"=IF(AA{i+1}>=Y{i+1},TRUE,FALSE)"
                    # isindf.at[i, 'Interest'] = = f"=AC{i+1}*D{i+1}/12"



                for tickers in isintickers:
                    tickers = list(tickers)
                    temptickerlist = []
                    if len(tickers) == 0:
                        isinrefassetlist.append('NoTickerFound')    
                        isinvallist.append(datetime.strftime(isindf['Observation Date'][tickcounter],'%m/%d/%Y'))
                        isinsettlelist.append(datetime.strftime(isindf['Settlement Date'][tickcounter],'%m/%d/%Y'))
                        isinlist.append(isindf['CUSIP'][tickcounter])
                        isinnotional.append(isindf['Notional'][tickcounter])
                        isininv.append(isindf['inventoryName'][tickcounter])
                        isinident.append(isindf['ident'][tickcounter])
                        isinfinaldaylist.append([datetime.strftime(isindf['Observation Date'][tickcounter],'%m/%d/%Y'),datetime.strftime(isindf['Settlement Date'][tickcounter],'%m/%d/%Y')\
                            ,isindf['CUSIP'][tickcounter],'NoTickerFound'])
                        isinauto.append(cusipdf['Autocall Field'][tickcounter])
                        isintpayingint.append(cusipdf['Paying Interest'][tickcounter])
                        isininterest.append(cusipdf['Interest'][tickcounter])
                        isinlongname.append(cusipdf['Long Name'][tickcounter])
                        
                        if ('memory' in isindf['Long Name'][tickcounter].lower()):
                            isinmemory.append('Refer to previous month')
                        else:
                            isinmemory.append('')

                    else:                  
                        for ticker in tickers:
                            if ticker != 'TD':
                                isinrefassetlist.append(ticker)
                                isinvallist.append(datetime.strftime(isindf['Observation Date'][tickcounter],'%m/%d/%Y'))
                                isinsettlelist.append(datetime.strftime(isindf['Settlement Date'][tickcounter],'%m/%d/%Y'))
                                isinlist.append(isindf['CUSIP'][tickcounter])
                                isinnotional.append(isindf['Notional'][tickcounter])
                                isininv.append(isindf['inventoryName'][tickcounter])
                                isinident.append(isindf['ident'][tickcounter])
                                isinfinaldaylist.append([datetime.strftime(isindf['Observation Date'][tickcounter],'%m/%d/%Y'),datetime.strftime(isindf['Settlement Date'][tickcounter],'%m/%d/%Y')\
                                    ,isindf['CUSIP'][tickcounter],ticker])
                        
                                isinauto.append(cusipdf['Autocall Field'][tickcounter])
                                isintpayingint.append(cusipdf['Paying Interest'][tickcounter])
                                isininterest.append(cusipdf['Interest'][tickcounter])
                                isinlongname.append(cusipdf['Long Name'][tickcounter])                            
                                
                                if ('memory' in isindf['Long Name'][tickcounter].lower()):
                                        isinmemory.append('Refer to previous month')
                                else:
                                    isinmemory.append('')                 
                                    
                    tickcounter += 1   
            
            # cusipdf.to_csv(outputdirectory+rf'\cusip_draft_'+rf'{datetime.strftime(day,"%b %#d")}.csv', index=False)
            # isindf.to_csv(outputdirectory+rf'\isin_draft_'+rf'{datetime.strftime(day,"%b %#d")}.csv', index=False)
            # df.to_csv(outputdirectory+rf'\df_'+rf'{datetime.strftime(day,"%b %#d")}.csv', index=False)

            wb = load_workbook(filename = fileloc)

            if cusipdf.empty == False:
                ws = wb['CUSIP']

                valdatecol = None
                paydatecol = None
                cusipcol = None
                cusipcolletter = None
                refassetcol = None
                principalcol = None
                notecol = None
                memorycol = None
                autocallcol = None
                autocolletter = None
                invcol = None
                identcol = None
                identcolletter = None
                payingintcol = None
                interestcol = None
                intcolletter = None


                def cellmaker(colletter,rownum,value):
                    ws[rf'{colletter}{rownum}'] = value
                    ws[rf'{colletter}{rownum}'].alignment = Alignment(horizontal='center',vertical='center')


                for i in range(len(cusiprefassetlist)):
                    #for row in ws.iter_rows(): # Need to optimize so it only goes through header
                        for cell in ws[1]:
                            if cell.value is not None:
                                if cell.value.strip() == 'Valuation Date':
                                    if valdatecol is None:
                                        valdatecol = cell.column
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipvallist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Payment Date':
                                    if paydatecol is None:
                                        paydatecol = cell.column
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipsettlelist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'ISIN / CUSIP':
                                    if cusipcol is None:
                                        cusipcol = cell.column
                                    if cusipcolletter is None:
                                        cusipcolletter = cell.column_letter                         
                                    ws[rf'{cell.column_letter}{i+2}'] = cusiplist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Reference Asset':
                                    if refassetcol is None:
                                        refassetcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = cusiprefassetlist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Principal':
                                    if principalcol is None:
                                        principalcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = abs(cusipnotional[i])
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'No. of Notes':
                                    if notecol is None:
                                        notecol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = abs(cusipnotional[i]/1000)
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Memory':
                                    if memorycol is None:
                                        memorycol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipmemory[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Inventory':
                                    if invcol is None:
                                        invcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipinv[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Ident':
                                    if identcol is None:
                                        identcol = cell.column
                                    if identcolletter is None:
                                        identcolletter = cell.column_letter                  
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipident[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Autocalled':
                                    if autocallcol is None:
                                        autocallcol = cell.column
                                    if autocolletter is None:
                                        autocolletter = cell.column_letter                            
                                    ws[rf'{cell.column_letter}{i+2}'] = cusipauto[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                   
                                if cell.value.strip() == 'Paying interest?':
                                    if payingintcol is None:
                                        payingintcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = cusiptpayingint[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                                                                 
                                if cell.value.strip() == 'Interest':
                                    if interestcol is None:
                                        interestcol = cell.column
                                    if intcolletter is None:
                                        intcolletter = cell.column_letter                                
                                    # ws[rf'{cell.column_letter}{i+2}'] = isininterest[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                                                                 
                            
                int_col = None
                ref_asset_col = None
                ident_col = None

                for col in range(1,ws.max_column+1):
                    if ws.cell(row=1,column=col).value == 'Interest':
                        int_col = col
                    elif ws.cell(row=1,column=col).value == 'Reference Asset':
                        ref_asset_col = col
                    elif ws.cell(row=1,column=col).value == 'Ident':
                        ident_col = col

                # fill out 'Interest' Column            
                for i in range(1,ws.max_row):
                    ws.cell(row=i+1, column=int_col).value = f"=AC{i+1}*D{i+1}/12"

                                                
                # Merging cells
                rows_iterate = []
                rowcounter = 2
                startmergerow = None
                for cell in ws[cusipcolletter]: # Only iterates through cusip column
                    if cell.row == 1:
                        continue
                    elif (cell.value is not None) & (cell.value != 1):
                        if ws[rf'{cell.column_letter}{rowcounter}'].value == ws[rf'{cell.column_letter}{rowcounter+1}'].value:
                            if startmergerow is not None:
                                rowcounter += 1
                                continue
                            else:
                                startmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                                rowcounter += 1
                                continue
                        else:
                            if startmergerow is None:
                                startmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                            endmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                            ws.merge_cells(start_row=startmergerow,start_column=cusipcol,end_row=endmergerow,end_column=cusipcol)
                            ws.merge_cells(start_row=startmergerow,start_column=valdatecol,end_row=endmergerow,end_column=valdatecol)
                            ws.merge_cells(start_row=startmergerow,start_column=paydatecol,end_row=endmergerow,end_column=paydatecol)
                            ws.merge_cells(start_row=startmergerow,start_column=principalcol,end_row=endmergerow,end_column=principalcol)
                            ws.merge_cells(start_row=startmergerow,start_column=notecol,end_row=endmergerow,end_column=notecol)
                            ws.merge_cells(start_row=startmergerow,start_column=memorycol,end_row=endmergerow,end_column=memorycol)
                            ws.merge_cells(start_row=startmergerow,start_column=invcol,end_row=endmergerow,end_column=invcol)
                            ws.merge_cells(start_row=startmergerow,start_column=int_col,end_row=endmergerow,end_column=int_col)
                            ws.merge_cells(start_row=startmergerow,start_column=autocallcol,end_row=endmergerow,end_column=autocallcol)
                            rows_iterate.append(startmergerow)
                            
                            # Sort cols     
                            RefAsset_Ident = []
                            
                            for row in range(startmergerow, endmergerow+1):
                                RefAsset_Ident.append([(ws.cell(row=row, column =ref_asset_col).value),(ws.cell(row=row, column =ident_col).value)])

                            # Sort the list first by "Ident" (the int value) then by "Reference Asset" (alphabetically)
                            RefAsset_Ident.sort(key=lambda x: (x[1],x[0]))
                            
                            for i in range(startmergerow, endmergerow+1):
                                ws.cell(row=i, column=ref_asset_col).value = RefAsset_Ident[i-startmergerow][0]
                                ws.cell(row=i, column=ident_col).value = RefAsset_Ident[i-startmergerow][1]

                            startmergerow = None
                            rowcounter += 1
                            
                # Merge "Ident" Column
                ident_iterate = []
                ident_counter = 2
                ident_start_merge_row = None
                for cell in ws[identcolletter]:
                    if cell.row == 1:
                        continue
                    elif (cell.value is not None) & (cell.value != 1):
                        if ws[rf'{cell.column_letter}{ident_counter}'].value == ws[rf'{cell.column_letter}{ident_counter+1}'].value:
                            if ident_start_merge_row is not None:
                                ident_counter += 1
                                continue
                            else:
                                ident_start_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                                ident_counter += 1                            
                                continue
                        else:
                            if ident_start_merge_row is None:
                                ident_start_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                            ident_end_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                            ws.merge_cells(start_row=ident_start_merge_row,start_column=identcol,end_row=ident_end_merge_row,end_column=identcol)
                            rows_iterate.append(ident_start_merge_row)
                            ident_start_merge_row = None
                            ident_counter += 1            
                

                # Change the font color to blue or red
                red = Font(color='FF0000', bold=True)
                blue =Font(color='0000FF')
                for d, h in zip(ws['AF'], ws['C']):
                    if d.row in rows_iterate:
                        if d.value == 'Valid IC Date':
                            d.font = red
                            h.font = blue
                    


            if isindf.empty == False:
                ws = wb['ISIN']

                valdatecol = None
                paydatecol = None
                cusipcol = None
                cusipcolletter = None
                refassetcol = None
                principalcol = None
                notecol = None
                memorycol = None
                autocallcol = None
                autocolletter = None
                invcol = None
                identcol = None
                identcolletter = None
                payingintcol = None
                interestcol = None
                intcolletter = None

                def cellmaker(colletter,rownum,value):
                    ws[rf'{colletter}{rownum}'] = value
                    ws[rf'{colletter}{rownum}'].alignment = Alignment(horizontal='center',vertical='center')


                for i in range(len(isinrefassetlist)):
                    #for row in ws.iter_rows(): # Need to optimize so it only goes through header
                        for cell in ws[1]:
                            if cell.value is not None:
                                if cell.value.strip() == 'Valuation Date':
                                    if valdatecol is None:
                                        valdatecol = cell.column
                                    ws[rf'{cell.column_letter}{i+2}'] = isinvallist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Payment Date':
                                    if paydatecol is None:
                                        paydatecol = cell.column
                                    ws[rf'{cell.column_letter}{i+2}'] = isinsettlelist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'ISIN / CUSIP':
                                    if cusipcol is None:
                                        cusipcol = cell.column
                                    if cusipcolletter is None:
                                        cusipcolletter = cell.column_letter                         
                                    ws[rf'{cell.column_letter}{i+2}'] = isinlist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Reference Asset':
                                    if refassetcol is None:
                                        refassetcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = isinrefassetlist[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Principal':
                                    if principalcol is None:
                                        principalcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = abs(isinnotional[i])
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'No. of Notes':
                                    if notecol is None:
                                        notecol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = abs(isinnotional[i]/1000)
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Memory':
                                    if memorycol is None:
                                        memorycol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = isinmemory[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Inventory':
                                    if invcol is None:
                                        invcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = isininv[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Ident':
                                    if identcol is None:
                                        identcol = cell.column
                                    if identcolletter is None:
                                        identcolletter = cell.column_letter                  
                                    ws[rf'{cell.column_letter}{i+2}'] = isinident[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                                if cell.value.strip() == 'Autocalled':
                                    if autocallcol is None:
                                        autocallcol = cell.column
                                    if autocolletter is None:
                                        autocolletter = cell.column_letter                            
                                    ws[rf'{cell.column_letter}{i+2}'] = isinauto[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                   
                                if cell.value.strip() == 'Paying interest?':
                                    if payingintcol is None:
                                        payingintcol = cell.column                            
                                    ws[rf'{cell.column_letter}{i+2}'] = isintpayingint[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                                                                 
                                if cell.value.strip() == 'Interest':
                                    if interestcol is None:
                                        interestcol = cell.column
                                    if intcolletter is None:
                                        intcolletter = cell.column_letter                                
                                    # ws[rf'{cell.column_letter}{i+2}'] = isininterest[i]
                                    ws[rf'{cell.column_letter}{i+2}'].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)                                                                 


                int_col = None
                ref_asset_col = None
                ident_col = None

                for col in range(1,ws.max_column+1):
                    if ws.cell(row=1,column=col).value == 'Interest':
                        int_col = col
                    elif ws.cell(row=1,column=col).value == 'Reference Asset':
                        ref_asset_col = col
                    elif ws.cell(row=1,column=col).value == 'Ident':
                        ident_col = col

                # fill out 'Interest' Column            
                for i in range(1,ws.max_row):
                    ws.cell(row=i+1, column=int_col).value = f"=AC{i+1}*D{i+1}/12"

                # Merging cells
                rowcounter = 2
                startmergerow = None
                for cell in ws[cusipcolletter]: # Only iterates through cusip column
                    if cell.row == 1:
                        continue
                    elif (cell.value is not None) & (cell.value != 1):
                        if ws[rf'{cell.column_letter}{rowcounter}'].value == ws[rf'{cell.column_letter}{rowcounter+1}'].value:
                            if startmergerow is not None:
                                rowcounter += 1
                                continue
                            else:
                                startmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                                rowcounter += 1
                                continue
                        else:
                            if startmergerow is None:
                                startmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                            endmergerow = ws[rf'{cell.column_letter}{rowcounter}'].row
                            ws.merge_cells(start_row=startmergerow,start_column=cusipcol,end_row=endmergerow,end_column=cusipcol)
                            ws.merge_cells(start_row=startmergerow,start_column=valdatecol,end_row=endmergerow,end_column=valdatecol)
                            ws.merge_cells(start_row=startmergerow,start_column=paydatecol,end_row=endmergerow,end_column=paydatecol)
                            ws.merge_cells(start_row=startmergerow,start_column=principalcol,end_row=endmergerow,end_column=principalcol)
                            ws.merge_cells(start_row=startmergerow,start_column=notecol,end_row=endmergerow,end_column=notecol)
                            ws.merge_cells(start_row=startmergerow,start_column=memorycol,end_row=endmergerow,end_column=memorycol)
                            ws.merge_cells(start_row=startmergerow,start_column=invcol,end_row=endmergerow,end_column=invcol)
                            ws.merge_cells(start_row=startmergerow,start_column=int_col,end_row=endmergerow,end_column=int_col)
                            ws.merge_cells(start_row=startmergerow,start_column=autocallcol,end_row=endmergerow,end_column=autocallcol)
                            rows_iterate.append(startmergerow)
                            
                            # Sort cols     
                            RefAsset_Ident = []
                            
                            for row in range(startmergerow, endmergerow+1):
                                RefAsset_Ident.append([(ws.cell(row=row, column =ref_asset_col).value),(ws.cell(row=row, column =ident_col).value)])

                            # Sort the list first by "Ident" (the int value) then by "Reference Asset" (alphabetically)
                            RefAsset_Ident.sort(key=lambda x: (x[1],x[0]))
                            
                            for i in range(startmergerow, endmergerow+1):
                                ws.cell(row=i, column=ref_asset_col).value = RefAsset_Ident[i-startmergerow][0]
                                ws.cell(row=i, column=ident_col).value = RefAsset_Ident[i-startmergerow][1]

                            startmergerow = None
                            rowcounter += 1
                            
                # Merge "Ident" Column
                ident_iterate = []
                ident_counter = 2
                ident_start_merge_row = None
                for cell in ws[identcolletter]:
                    if cell.row == 1:
                        continue
                    elif (cell.value is not None) & (cell.value != 1):
                        if ws[rf'{cell.column_letter}{ident_counter}'].value == ws[rf'{cell.column_letter}{ident_counter+1}'].value:
                            if ident_start_merge_row is not None:
                                ident_counter += 1
                                continue
                            else:
                                ident_start_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                                ident_counter += 1                            
                                continue
                        else:
                            if ident_start_merge_row is None:
                                ident_start_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                            ident_end_merge_row = ws[rf'{cell.column_letter}{ident_counter}'].row
                            ws.merge_cells(start_row=ident_start_merge_row,start_column=identcol,end_row=ident_end_merge_row,end_column=identcol)
                            rows_iterate.append(ident_start_merge_row)
                            ident_start_merge_row = None
                            ident_counter += 1            
                

                # Change the font color to blue or red
                red = Font(color='FF0000', bold=True)
                blue =Font(color='0000FF')
                for d, h in zip(ws['AF'], ws['C']):
                    if d.row in rows_iterate:
                        if d.value == 'Valid IC Date':
                            d.font = red
                            h.font = blue
            
                            # for col in range(1,ws.max_column+1):
                            #     if ws.cell(row=1, column=col).value == "Reference Asset":
                            #         ref_asset_col = col
                            #         break
                            
                            # ref_asset_values = []
                            # for row in range(startmergerow, endmergerow+1):
                            #     ref_asset_values.append((ws.cell(row=row, column =ref_asset_col).value))
                                                                        
                            # ref_asset_values.sort(key=lambda x: x[0])
                            
                            # for i in range(startmergerow, endmergerow+1):
                            #     ws.cell(row=i, column=ref_asset_col).value = ref_asset_values[i-startmergerow]   
                                                
                            # startmergerow = None
                            # rowcounter += 1

            wb.save(fileloc)
            wb.close()
    
    testcounter += 1

    
